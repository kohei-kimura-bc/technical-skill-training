import openpyxl
wb = openpyxl.load_workbook("test.xlsx") # Excelファイルの読み込み
print(f"コピー前のsheet名：　{wb.sheetnames}")
ws = wb.copy_worksheet(wb["test2"])
print(f"コピー後のsheet名：　{wb.sheetnames}") #test2 Copyという名前のsheetが追加されている

wb.save("copy_worksheet.xlsx") #test2 Copyという名前のsheetが追加されたファイルが保存される