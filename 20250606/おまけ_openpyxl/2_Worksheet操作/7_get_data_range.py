import openpyxl
wb = openpyxl.load_workbook("test.xlsx") # Excelファイルの読み込み

#test1シートについて
ws = wb["test1"] # シートの取得

min_row = ws.min_row # 行の最小
max_row = ws.max_row # 行の最大
min_col = ws.min_column # 列の最小
max_col = ws.max_column # 列の最大

print("test1シートのデータ記載範囲は")
print(F"行の最小{min_row}・行の最大{max_row}")
print(F"列の最小{min_col}・列の最大{max_col}\n")


#test2シートについて
ws = wb["test2"] # シートの取得

min_row = ws.min_row # 行の最小
max_row = ws.max_row # 行の最大
min_col = ws.min_column # 列の最小
max_col = ws.max_column # 列の最大

print("test2シートのデータ記載範囲は")
print(F"行の最小{min_row}・行の最大{max_row}")
print(F"列の最小{min_col}・列の最大{max_col}")