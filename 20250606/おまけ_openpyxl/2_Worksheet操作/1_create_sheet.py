import openpyxl
wb = openpyxl.load_workbook("test.xlsx") # Excelファイルの読み込み
ws = wb.create_sheet(title="追加シート") # シートの追加

wb.save("create_sheet.xlsx") #1番右にシートが追加されたファイルが保存される


#追加位置も指定
import openpyxl
wb = openpyxl.load_workbook("test.xlsx") # Excelファイルの読み込み
ws = wb.create_sheet(title="追加シート", index=0) # シートの追加

wb.save("create_sheet_with_index.xlsx") #index=0としたため1番左にシートが追加されたファイルが保存される